import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

class fdc:
    def __init__(self,file):
        self.fdate=dict()
        self.file=file

    def set_fd_date(self,day,month,year,fg):
        self.fdate[fg]=datetime.date(year,month,day)

    def get_fd_date(self,fg):
        day=str(self.fdate[fg].day)
        month=str(self.fdate[fg].month)
        year=str(self.fdate[fg].year)
        print(day+"/"+month+"/"+year)

    def set_value(self,accno,amount,roi,owner='Madu',bank='hdfc'):
        self.accno = accno
        self.amount = amount
        self.roi = roi
        self.owner=owner
        self.bank=bank

    def save(self):
        wb = load_workbook(self.file)
        ws = wb.worksheets[0]
        cdate=datetime.datetime.now()
        edate=self.fdate["end"]
        if(edate>cdate):
            self.active=1
        else:
            self.active=0
        data1 = [self.accno,self.amount,self.roi,self.fdate["start"],self.fdate["end"],self.owner,self.bank,self.active]
        ws.append(data1)
        wb.save(self.file)

    def fetch_fd_value_all(self):
        amlist=[]
        for i in self.fetch_acc_list():
            amlist.append(self.fetch_fd_value(i))
        return(amlist)

    def fetch_fd_value_all_filter(self,filter,field):
        amlist=[]
        for i in self.fetch_acc_list_filter(filter,field):
            amlist.append(self.fetch_fd_value(i))
        return(amlist)

    def fetch_fd_value_all_filter_date(self,filter,field,date1,date2):
        amlist=[]
        for i in self.fetch_acc_list_filter(filter,field):
            amlist.append(self.fetch_fd_value(i))
        return(amlist)

    def fetch_acc_list_filter(self,field,fvalue):
        wb = load_workbook(self.file)
        ws = wb.worksheets[0]
        alist=[]
        j=0
        for i in ws.rows:
            for k in i:
                if(k.value==field):
                    break
                j=j+1
            break
        for i in ws.rows:
            if(i[j].value!=field):
                if(i[j].value==fvalue):
                    alist.append(i[0].value)
        return(alist)

    def fetch_acc_list(self):
        wb = load_workbook(self.file)
        ws = wb.worksheets[0]
        alist=[]
        for i in ws.rows:
            if(i[0].value!='Acc no'):
                alist.append(i[0].value)
        return(alist)

    def fetch_fd_value_date(self, accno,date1,date2,fg):
        wb = load_workbook(self.file)
        ws = wb.worksheets[0]
        for i in ws.rows:
            if(i[0].value==accno):
                amount=i[1].value
                roi=i[2].value
                sdate=i[3].value
                edate=i[4].value
                break
        if((date1 > sdate)and(edate>date1)):
            sdate=date1
        if ((date2 < edate)and(date2>sdate)):
            edate = date2
        rdate = relativedelta(edate,sdate)
        fhlist=[]
        if(fg==0):
            m=(rdate.years*12)+rdate.months+1
            for i in range(m):
                cdate=sdate+relativedelta(months=+i)
                fhlist.append(self.fetch_fd_value(accno,cdate=cdate))
        else:
            y=rdate.years+1
            for i in range(y):
                cdate = sdate + relativedelta(years=+i)
                fhlist.append(self.fetch_fd_value(accno, cdate=cdate))
        print(fhlist)

    def fetch_fd_value(self, accno, cdate=datetime.datetime.now()):
        wb = load_workbook(self.file)
        ws = wb.worksheets[0]
        for i in ws.rows:
            if(i[0].value==accno):
                amount=i[1].value
                roi=i[2].value
                sdate=i[3].value
                edate=i[4].value
                break
        if(cdate>edate):
            rdelta = relativedelta(edate, sdate)
            interest = (int(rdelta.years) * roi * amount) / 100
            interest = interest + ((rdelta.months * roi * amount) / 1200)
        else:
            rdelta = relativedelta(cdate, sdate)
            interest=(int(rdelta.years)*roi*amount)/100
            interest=interest+((rdelta.months*roi*amount)/1200)
        return(int(interest)+amount)



